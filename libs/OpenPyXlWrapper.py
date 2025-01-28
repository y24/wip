from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def load(file_path:str):
    return load_workbook(file_path)


def get_sheet_by_name(workbook, sheet_name:str):
    return workbook[sheet_name]


def load_sheet_as_array(sheet, header_row=1, fill_blank=False):
    # 1行目の値をキーとして取得（空白セルがある列は無視）
    keys = [cell.value for cell in sheet[header_row]]
    valid_indices = [i for i, key in enumerate(keys) if key is not None]
    keys = [keys[i] for i in valid_indices]

    # データを格納するリスト
    data = []

    # 前の行の値を保持するリスト
    previous_row = [None] * len(keys)

    try:
        # 2行目以降を辞書として取得
        for row in sheet.iter_rows(min_row=header_row + 1, max_col=len(keys), values_only=True):
            if fill_blank:
                # 空白セルを前の行の値で埋める
                row_filled = []
                for i, index in enumerate(valid_indices):
                    try:
                        value = row[index]
                    except IndexError:
                        raise ValueError("指定されたヘッダー列以降にデータがありません。")
                    if value is None:
                        value = previous_row[i]  # 前の行の値を使用
                    row_filled.append(value)
                row = row_filled
            
            # 行が全て空白の場合はスキップ
            if all(value is None for value in row):
                continue
            
            # 辞書を生成して追加
            data.append(row)

            # 現在の行を前の行として保存
            previous_row = row

        # データなし
        if not data:
            raise ValueError("取得できるデータがありません。")

        return data
    except ValueError as e:
        print(f"Error: {e}")


def load_sheet_as_dictionary(sheet, header_row=1, fill_blank=False):
    # 1行目の値をキーとして取得（空白セルがある列は無視）
    keys = [cell.value for cell in sheet[header_row]]
    valid_indices = [i for i, key in enumerate(keys) if key is not None]
    keys = [keys[i] for i in valid_indices]

    # データを格納するリスト
    data = []

    # 前の行の値を保持するリスト
    previous_row = [None] * len(keys)

    try:
        # 2行目以降を取得
        for row in sheet.iter_rows(min_row=header_row + 1, max_col=len(keys), values_only=True):
            if fill_blank:
                # 空白セルを前の行の値で埋める
                row_filled = []
                for i, index in enumerate(valid_indices):
                    try:
                        value = row[index]
                    except IndexError:
                        raise ValueError("指定されたヘッダー列以降にデータがありません。")
                    if value is None:
                        value = previous_row[i]  # 前の行の値を使用
                    row_filled.append(value)
                row = row_filled
            
            # 行が全て空白の場合はスキップ
            if all(value is None for value in row):
                continue
            
            # 辞書を生成して追加
            data.append(row)

            # 現在の行を前の行として保存
            previous_row = row

        # データなし
        if not data:
            raise ValueError("取得できるデータがありません。")

        return data
    except ValueError as e:
        print(f"Error: {e}")


def find_row(sheet, search_row:str, search_str:str):
    try:
        # 列名
        col_num = column_index_from_string(search_row) - 1

        # 指定列をループして値を確認
        for row in sheet.iter_rows(min_col=1, max_col=1):
            cell = row[col_num]  # 指定列のセル
            if cell.value == search_str:  # 値が search_str のセル
                return cell.row
        return None

    except Exception as e:
        print(f"Error: {e}")